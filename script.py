import streamlit as st
import pandas as pd
import io
from streamlit_pdf_viewer import pdf_viewer
import re
import json
from io import StringIO, BytesIO
import base64
from openpyxl import load_workbook
import os
import warnings
import streamlit.components.v1 as components

ruta_plantilla = "Plantilla_Turbo_Final.xlsx"


def obtener_tabla_habitaciones():
    if "costos_excel" in st.session_state:
        df_costos = st.session_state["costos_excel"].copy()

        # Filtrar solo las columnas necesarias
        columnas_exportar = [
            "Item",
            "ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS",
            "Unidad",
            "Valor Unitario ofertado (**)"
        ]
        df_intermedio = df_costos[columnas_exportar].copy()

        # -----------------------------------------------
        # 1. Crear la columna 'Categoria' replicando la lógica de "categorias_actividades"
        # -----------------------------------------------
        categoria_actual = None
        categorias = []
        
        for _, row in df_intermedio.iterrows():
            actividad = str(row["ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS"])
            # Si la actividad es mayúscula, asumimos que es una nueva categoría
            if actividad.isupper():
                categoria_actual = actividad
                categorias.append(categoria_actual)
            else:
                # No es mayúscula, así que sigue perteneciendo a la última categoría
                categorias.append(categoria_actual)
        
        df_intermedio["Categoria"] = categorias

        # 2. Para cada habitación procesada, creamos una columna con la cantidad usada
        if "resultados_csv" in st.session_state:
            habitaciones_procesadas = [
                habitacion
                for habitacion in st.session_state["resultados_csv"].keys()
                if "piso" not in habitacion.lower()
            ]

            for habitacion in habitaciones_procesadas:
                df_intermedio[habitacion] = 0.0
                for i, row in df_intermedio.iterrows():
                    actividad = row["ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS"]
                    cantidad_key = f"cantidad_{habitacion}_{actividad}"
                    if cantidad_key in st.session_state:
                        df_intermedio.at[i, habitacion] = st.session_state[cantidad_key]

            # 3. Sumar las columnas de habitaciones para obtener 'Total actividad'
            df_intermedio["Total actividad"] = df_intermedio[habitaciones_procesadas].sum(axis=1)

            # 4. Costo total
            df_intermedio["Costo total"] = (
                df_intermedio["Total actividad"] *
                df_intermedio["Valor Unitario ofertado (**)"]
            )

            # 5. Crear DataFrame resumen (ahora con la columna 'Categoria')
            df_resumen = df_intermedio[[
                "Item",
                "Categoria",
                "ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS",
                "Unidad",
                "Valor Unitario ofertado (**)",
                "Total actividad",
                "Costo total"
            ]].copy()

        # 6. Generar el archivo Excel con la plantilla
        nueva_ruta = export_to_excel(df_resumen)
        st.session_state["export_excel"] = nueva_ruta


def export_to_excel(df_summary):
    """
    Llena la plantilla con las actividades > 0,
    agrupándolas por la columna 'Categoria' (encabezado en mayúsculas).
    Formatea las columnas M, N y O como moneda sin decimales; 
    además, el valor de la columna N se replica en la columna O (Subtotal).
    Finalmente, se realiza una autosuma de la columna O (desde la fila 31 a la 93)
    y se almacena en la celda O94.
    """
    ruta_plantilla = os.path.join(os.getcwd(), "Plantilla_Turbo_Final.xlsx")
    if not os.path.exists(ruta_plantilla):
        st.error(f"⚠️ No se encontró la plantilla: {ruta_plantilla}")
        return None

    wb = load_workbook(ruta_plantilla)
    ws = wb.active  # Hoja principal de la plantilla

    # Obtener celdas combinadas para evitar sobreescritura
    celdas_combinadas = set()
    for merged_range in ws.merged_cells.ranges:
        for row in ws[merged_range.coord]:
            for cell in row:
                celdas_combinadas.add(cell.coordinate)

    # 1. Filtrar solo actividades con Total actividad > 0
    df_filtrado = df_summary[df_summary["Total actividad"] > 0].copy()

    # 2. Si no hay filas, guardar y salir
    if df_filtrado.empty:
        st.warning("No hay actividades con valor > 0. El Excel quedará vacío.")
        nueva_ruta = os.path.join(os.getcwd(), "Reporte_Resultado.xlsx")
        wb.save(nueva_ruta)
        return nueva_ruta

    # 3. Obtener el orden de las categorías tal como aparecen en df_filtrado
    categorias_unicas = list(df_filtrado["Categoria"].dropna().unique())

    # 4. Empezar a escribir en la fila 31 (espacio reservado para encabezados, etc.)
    current_row = 31

    for cat in categorias_unicas:
        if not cat:  # Omitir categorías vacías
            continue

        # Tomar las actividades de esta categoría
        df_cat = df_filtrado[df_filtrado["Categoria"] == cat]
        if df_cat.empty:
            continue

        # Escribir el título de la categoría en la columna A
        if f"A{current_row}" not in celdas_combinadas:
            ws[f"A{current_row}"] = cat
        current_row += 1

        # Para cada actividad, definir el mapeo de columnas y valores
        # Nuevo orden: A, B, K, M, L, N, O, G
        for _, fila in df_cat.iterrows():
            col_map = ["A", "B", "K", "M", "L", "N", "O", "G"]
            # Se toma "Costo total" para columna N y se replica para columna O.
            valores = [
                fila["Item"],
                fila["ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS"],
                fila["Unidad"],
                fila["Valor Unitario ofertado (**)"],
                fila["Total actividad"],
                fila["Costo total"],
                fila["Costo total"],  # Replicado en columna O (Subtotal)
                ""  # Puedes dejar en blanco o agregar otra información en la columna G
            ]

            for col, val in zip(col_map, valores):
                celda = f"{col}{current_row}"
                if celda not in celdas_combinadas:
                    # Si la columna es M, N o O, redondeamos el valor (sin decimales) y aplicamos formato de moneda
                    if col in ["M", "N", "O"]:
                        try:
                            valor_num = int(round(float(val)))
                        except Exception as e:
                            valor_num = 0
                        ws[celda].value = valor_num
                        ws[celda].number_format = '"$"#,##0'
                    else:
                        ws[celda].value = val
            current_row += 1

        # Fila en blanco entre categorías
        current_row += 1

    # Realizar la autosuma en la columna O y colocarla en la fila 94.
    # Se suma desde la fila 31 hasta la 93 (ajusta el rango si es necesario)
    ws["O94"] = "=SUM(O31:O93)"
    ws["O94"].number_format = '"$"#,##0'

    # Guardar el archivo resultante
    nueva_ruta = os.path.join(os.getcwd(), "Reporte_Resultado.xlsx")
    wb.save(nueva_ruta)

    return nueva_ruta

def procesar_csv_bytes(file_bytes: BytesIO):
    """
    Procesa un archivo CSV desde un BytesIO y devuelve un diccionario con las tablas encontradas.

    Args:
        file_bytes (BytesIO): Archivo CSV en memoria.

    Returns:
        tuple: Un diccionario con las tablas y un código de estado HTTP.
    """
    try:
        content = file_bytes.getvalue().decode('utf-8', errors='replace')

        raw_sections = re.split(r'\n\s*\n+', content)
        sections = [sec.strip() for sec in raw_sections if sec.strip()]
        
        tablas = {}
        for idx, section in enumerate(sections, start=1):
            lines = section.split('\n')

            if len(lines) == 1:
                tablas[f"tabla_{idx}"] = {"titulo": lines[0]}
                continue
            
            if all(':' in line for line in lines if line.strip()):
                data = {key.strip(): value.strip().strip(',')
                        for line in lines if (parts := line.split(':', 1)) and len(parts) == 2
                        for key, value in [parts]}
                tablas[f"tabla_{idx}"] = data
                continue
            
            try:
                read_csv_kwargs = {"encoding": "utf-8"}
                if pd.__version__ >= "1.3.0":
                    read_csv_kwargs["on_bad_lines"] = "skip"
                else:
                    read_csv_kwargs["error_bad_lines"] = False
                
                df = pd.read_csv(StringIO(section), **read_csv_kwargs)
                
                if not df.empty:
                    df.columns = df.columns.str.strip()
                    tablas[f"tabla_{idx}"] = df
                    continue
            except pd.errors.ParserError:
                pass  

            data = {f"columna_{i}": [part.strip() for part in line.split(',')] 
                    if ',' in line else line.strip() for i, line in enumerate(lines)}
            tablas[f"tabla_{idx}"] = data

        return tablas, 200
    except UnicodeDecodeError:
        return {"error": "Error al leer el archivo, posible problema de codificación"}, 400
    except Exception as e:
        return {"error": f"Error al procesar el archivo CSV: {str(e)}"}, 500

def calcular_propiedades_habitacion(tablas):
    """
    Calcula valores para cada habitación en las tablas encontradas.

    Args:
        tablas (dict): Diccionario de tablas procesadas.

    Returns:
        dict: JSON con los resultados en formato de diccionario.
    """
    resultados = {}

    for tabla_key, value in tablas.items():
        if isinstance(value, pd.DataFrame):
            df = value.copy()
            df.columns = df.columns.str.strip()

            columnas_requeridas = ["Tierra Superficie: : m²", "Paredes sin apertura: m²"]
            if not all(col in df.columns for col in columnas_requeridas):
                continue

            for _, row in df.iterrows():
                try:
                    nombre_habitacion = row.iloc[0]  # Primera columna es el nombre

                    superficie = float(row.get("Tierra Superficie: : m²", 0) or 0)
                    paredes_sin_apertura = float(row.get("Paredes sin apertura: m²", 0) or 0)
                    perimetro_interno = float(row.get("Tierra Perímetro: m", 0) or 0)
                    perimetro_techo = float(row.get("Techo Perímetro: m", 0) or 0)
                    diferencia = abs(perimetro_interno - perimetro_techo)
                    techo = superficie * 1.15 if diferencia >= 0.1 else superficie

                    resultados[nombre_habitacion] = {
                        "MAGICPLAN - ÁREA PISO": superficie,
                        "MAGICPLAN - ÁREA PARED": paredes_sin_apertura,
                        "MAGICPLAN - ÁREA CUBIERTA": techo,
                        "MAGICPLAN - PERIMETRO PISO": perimetro_interno,
                        "MAGICPLAN - PERIMETRO CUBIERTA": perimetro_techo,
                    }
                    
                except Exception as e:
                    resultados[f"Error en {tabla_key}"] = f"Error al procesar habitación: {str(e)}"

    return resultados

@st.cache_data
def load_pdf(file):
    return file.read()

@st.cache_data
def load_image(file):
    return Image.open(file)

def inicio():
    
    
    
    st.title("Ingreso de archivos")
    st.write("Cargue los archivos correspondientes a la vivienda.")

    # Carga automática del archivo Excel sin necesidad de subirlo manualmente
    try:
        st.session_state["costos_excel"] = load_excel_local()
        st.success("Archivo Excel de costos cargado correctamente desde el código.")
    except Exception as e:
        st.error(f"Error al cargar el archivo Excel: {str(e)}")

    # Cargar archivos desde la interfaz web
    plano_file = st.file_uploader("Sube un archivo (Plano o Imagen)", type=["pdf", "png", "jpg", "jpeg"])
    resultados_csv = st.file_uploader("Sube un archivo CSV (Resultados MagicPlan)", type=["csv"])

    # Validar que ambos archivos sean subidos antes de continuar
    if plano_file and resultados_csv:
        file_extension = plano_file.name.split(".")[-1].lower()

        # Si es un PDF, lo carga con load_pdf()
        if file_extension == "pdf":
            st.session_state["plano_pdf"] = load_pdf(plano_file)
            st.success("Archivo PDF cargado correctamente.")

        # Si es una imagen (PNG, JPG, JPEG), la carga con load_image()
        elif file_extension in ["png", "jpg", "jpeg"]:
            st.session_state["plano_img"] = load_image(plano_file)
            st.success("Imagen cargada correctamente.")

        # Procesar el archivo CSV como antes
        tablas, codigo = procesar_csv_bytes(resultados_csv)
        st.session_state["resultados_csv"] = calcular_propiedades_habitacion(tablas)

        st.success("Todos los archivos han sido cargados correctamente.")

@st.cache_data
def load_image(file):
    return Image.open(file)

def detectar_desconexion():
    """Inserta un script en JavaScript para detectar desconexión y reconectar automáticamente"""
    html_code = """
    <script>
    function checkConnection() {
        if (!navigator.onLine) {
            alert("¡Se ha perdido la conexión! La aplicación seguirá funcionando.");
        }
    }
    setInterval(checkConnection, 5000); // Verifica la conexión cada 5 segundos
    </script>
    """
    components.html(html_code, height=0)


def main():
    
    st.set_page_config(page_title="Modificación de vivienda", layout="wide")
    
    detectar_desconexion()  # Agregar detección de pérdida de conexión
    
    if st.sidebar.button("Reiniciar aplicación"):
        st.session_state.clear()  # Limpia todos los valores almacenados
        st.rerun()
    
    # 🔹 Valor máximo permitido fijo
    max_total = 15600000  # 15.600.000

    # 🔹 Restar automáticamente 1.300.000 para obtener el diagnóstico
    diagnostico = max_total - 1300000  # 15.600.000 - 1.300.000

    # 📌 Mostrar ambos valores en la barra lateral
    st.sidebar.markdown(f"**Valor máximo permitido: ${max_total:,.2f}**")
    st.sidebar.markdown(f"**Valor con DIAGNÓSTICO: ${diagnostico:,.2f}** 🏥")

    # 🔹 El usuario aún puede reducir el costo con un porcentaje
    max_porcentaje = st.sidebar.number_input(
        "Ingrese el porcentaje de costos a reducir", 
        min_value=0.0, 
        max_value=100.0, 
        format="%.1f", 
        step=0.1, 
        key="max_porcentaje"
    )

    # 🔹 Calcular el nuevo costo permitido después de la reducción
    st.session_state['max_costo'] = diagnostico * (100 - max_porcentaje) / 100

    # 📌 Mostrar el valor final después de la reducción
    st.sidebar.markdown(f"**Costo permitido después de reducción: ${st.session_state['max_costo']:,.2f}**")

    # 🔹 Continuar con las pantallas de la aplicación
    inicio()
    vista_archivos(st.session_state['max_costo'])

@st.cache_data
def load_pdf(file):
    return file.read()

@st.cache_data
def load_csv(file):
    return pd.read_csv(file)

# Ruta del archivo Excel local (ajusta esto a tu ubicación real)
RUTA_ARCHIVO_COSTOS = "TURBO_ARCHIVO_PARA_TRABAJAR.xlsx"

# Función para cargar el archivo Excel desde la ruta local
@st.cache_data
def load_excel_local():
    return pd.read_excel(RUTA_ARCHIVO_COSTOS, sheet_name="FORMATO DE OFERTA ECONÓMICA")


def ultimas_dos_palabras(texto: str) -> str:
    palabras = texto.split()  # Dividir el texto en palabras
    return " ".join(palabras[-2:]) if len(palabras) >= 2 else texto

def verificar_palabras(texto, lista_referencia):
    palabras = {palabra.strip() for palabra in texto.split(",")}  # Convertir en conjunto sin espacios extra
    lista_referencia = set(lista_referencia)  # Convertir la lista en conjunto
    return not palabras.isdisjoint(lista_referencia)

def vista_archivos(max_total):
    st.title("Modificaciones a realizar")

    # Mostrar el archivo PDF si se ha subido
    if "plano_pdf" in st.session_state:
        st.subheader("Plano PDF")
        pdf_viewer(st.session_state["plano_pdf"], width="50%")

    # Mostrar la imagen si el usuario subió una imagen en lugar de un PDF
    elif "plano_img" in st.session_state:
        st.subheader("Plano en Imagen")
        st.image(st.session_state["plano_img"], caption="Plano en imagen", use_container_width=True)

    # Si los archivos CSV y Excel están cargados, mostrar la interfaz de modificaciones
    if "resultados_csv" in st.session_state and "costos_excel" in st.session_state:
        st.subheader("Selección de Habitaciones")
        habitaciones = [key for key in st.session_state["resultados_csv"].keys() if "piso" not in key.lower()]
        actividades = st.session_state["costos_excel"]
        estados = {}
        subtotales = {}

        for habitacion in habitaciones:
            activo = habitacion.startswith("#")
            estados[habitacion] = st.checkbox(habitacion, value=activo, key=f"habitacion_{habitacion}")
            subtotal = 0.0

            if estados[habitacion]:
                st.subheader(f"🏠 Modificaciones de {habitacion}")  # Quitamos el expander de habitación

                # Crear un diccionario para almacenar categorías con actividades
                categorias_actividades = {}
                categoria_actual = None

                for _, row in actividades.iterrows():
                    actividad = row.get("ACTIVIDAD DE OBRA - LISTA DE PRECIOS UNITARIOS", "")
                    unidad = row.get("Unidad", None)
                    item = row.get("Item", "")
                    valor_unitario = row.get("Valor Unitario ofertado (**)", 0.0)
                    medicion = row.get("ÁREA", "")
                    formula = row.get("FORMULA", "")
                    formula = "" if pd.isna(formula) else formula

                    # Identificar si el título es una categoría (mayúsculas)
                    if actividad.isupper():
                        categoria_actual = actividad
                        categorias_actividades[categoria_actual] = []
                    elif categoria_actual:
                        categorias_actividades[categoria_actual].append((item, actividad, unidad, valor_unitario, medicion, formula))

                # Mostrar las categorías con sus actividades dentro de `st.expander()`
                for categoria, lista_actividades in categorias_actividades.items():
                    with st.expander(f"📂 {categoria}", expanded=False):
                        for item, actividad, unidad, valor_unitario, medicion, formula in lista_actividades:
                            check = st.checkbox(f"{item} -- {actividad} [Unidad: {unidad}] (Precio unitario: ${valor_unitario:,.2f})", key=f"check_{habitacion}_{actividad}")

                            if check:
                                cantidad_key = f"cantidad_{habitacion}_{actividad}"
                                valor_guardado_key = f"valor_{habitacion}_{actividad}"
                                if valor_guardado_key not in st.session_state:
                                    st.session_state[valor_guardado_key] = 0.0
                                if "USUARIO" in medicion.upper():
                                    cantidad = st.number_input(f"Ingrese la cantidad ({unidad}).", min_value=0 if unidad in ["UN", "UND"] else 0.00, key=cantidad_key, step=1 if unidad in ["UN", "UND"] else 0.0001)
                                    if st.button(f"Guardar cantidad", key=f"button_{habitacion}_{actividad}"):
                                        st.session_state[valor_guardado_key] = cantidad * valor_unitario
                                        st.success(f"Valor guardado para {actividad}: ${st.session_state[valor_guardado_key]:,.2f}")
                                else:
                                    if "ALTURA" in formula:    
                                        cantidad = st.number_input(f"Valor MagicPlan ({ultimas_dos_palabras(medicion)})", value=st.session_state["resultados_csv"][habitacion][medicion], min_value=0.0, key=cantidad_key)
                                        valor_input = st.number_input(f"Ingrese la altura (metros).", min_value=0.00, key=cantidad_key+"_aux")
                                        if st.button(f"Guardar cantidad", key=f"button_{habitacion}_{actividad}"):
                                            st.session_state[valor_guardado_key] = cantidad * valor_unitario * valor_input
                                            st.success(f"Valor guardado para {actividad}: ${st.session_state[valor_guardado_key]:,.2f}")
                                    elif formula != "":
                                        cantidad = st.number_input(f"Ingrese la cantidad ({unidad}).", value=st.session_state["resultados_csv"][habitacion][medicion], min_value=0.0, key=cantidad_key)
                                        if st.button(f"Guardar cantidad", key=f"button_{habitacion}_{actividad}"):
                                            st.session_state[valor_guardado_key] = cantidad * valor_unitario
                                            st.success(f"Valor guardado para {actividad}: ${st.session_state[valor_guardado_key]:,.2f}")
                                    else:    
                                        cantidad = st.number_input(f"Valor MagicPlan ({ultimas_dos_palabras(medicion)}) [Unidad: {unidad}]", value=st.session_state["resultados_csv"][habitacion][medicion], min_value=0.0, key=cantidad_key)
                                        st.session_state[valor_guardado_key] = cantidad * valor_unitario
                                        st.success(f"Valor guardado para {actividad}: ${st.session_state[valor_guardado_key]:,.2f}")
                                subtotal += st.session_state[valor_guardado_key]

            subtotales[habitacion] = subtotal

        total_general = sum(subtotales.values())
        st.sidebar.subheader("Subtotales por Habitación")
        df_subtotales = pd.DataFrame(list(subtotales.items()), columns=["Habitación", "Subtotal ($)"]).round(2)
        st.sidebar.dataframe(df_subtotales, hide_index=True)
        st.sidebar.subheader("Total General")

        if total_general > max_total:
            st.sidebar.markdown(f"<span style='color: red; font-weight: bold;'>Total: ${total_general:,.2f}</span>", unsafe_allow_html=True)
            st.sidebar.warning('Se ha superado el monto máximo permisible.')
        else:
            st.sidebar.markdown(f"Total: ${total_general:,.2f}")
            obtener_tabla_habitaciones()
            
            # 🔹 MODIFICACIÓN: Descargar el archivo Excel generado con la plantilla
            if "export_excel" in st.session_state and total_general > 0:
                try:
                    with open(st.session_state["export_excel"], "rb") as file:
                        st.sidebar.download_button(
                            label="Descargar Reporte",
                            data=file,
                            file_name="Reporte_Resultado.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                except Exception as e:
                    st.sidebar.error(f"Error al generar el archivo: {str(e)}")
    else:
        st.warning('Ingrese los archivos para iniciar el proceso, en la sección Inicio.')
        
def registro_login():
    st.title("Registro o Inicio de Sesión")
    opcion = st.radio("Elige una opción:", ["Iniciar Sesión", "Registrarse"])
    
    if opcion == "Iniciar Sesión":
        usuario = st.text_input("Usuario")
        contraseña = st.text_input("Contraseña", type="password")
        if st.button("Ingresar"):
            st.success(f"Bienvenido, {usuario}!")
    
    elif opcion == "Registrarse":
        nuevo_usuario = st.text_input("Nuevo Usuario")
        nueva_contraseña = st.text_input("Nueva Contraseña", type="password")
        confirmar_contraseña = st.text_input("Confirmar Contraseña", type="password")
        if st.button("Registrarse"):
            if nueva_contraseña == confirmar_contraseña:
                st.success("Registro exitoso. Ahora puedes iniciar sesión.")
            else:
                st.error("Las contraseñas no coinciden.")

if __name__ == "__main__":
    main()
